"""
Read sektorler.xlsx and analyze sector distribution
"""
import openpyxl
import sys
import io

# Fix encoding for Windows console
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook('sektorler.xlsx')
ws = wb.active

print("IS YATIRIM SEKTOR LISTESI ANALIZI")
print("=" * 100)

# Get headers
headers = [cell.value for cell in ws[1]]
print(f"\nSütunlar: {headers}")

# Read all data
data = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0]:  # If ticker exists
        data.append(row)

print(f"\nToplam şirket: {len(data)}")

# Count unique sectors - GERÇEK SEKTÖR 3. SÜTUNDA (index 2)
sectors = {}
for row in data:
    sector = row[2] if len(row) > 2 and row[2] else 'N/A'
    if sector not in sectors:
        sectors[sector] = []
    sectors[sector].append((row[0], row[1]))

print(f"\nBenzersiz sektör sayısı: {len(sectors)}")
print("\nSektör dağılımı:")
print("-" * 100)

for sector, companies in sorted(sectors.items(), key=lambda x: len(x[1]), reverse=True):
    print(f"{sector:50} {len(companies):3} şirket")
    # Show first 5 companies
    if len(companies) <= 5:
        for ticker, name in companies:
            print(f"    • {ticker:8} {name}")
    else:
        for ticker, name in companies[:5]:
            print(f"    • {ticker:8} {name}")
        print(f"    ... ve {len(companies)-5} şirket daha")

print("\n" + "=" * 100)
print("\nİlk 20 şirket (örnek):")
print("-" * 100)

for i, row in enumerate(data[:20], 1):
    ticker = row[0]
    sector = row[1] if len(row) > 1 else 'N/A'
    name = row[2] if len(row) > 2 else 'N/A'
    print(f"{i:2}. {ticker:8} {sector:40} {name}")

wb.close()
